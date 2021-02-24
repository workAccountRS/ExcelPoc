import cx_Oracle

import config
from ExcelHandler import ExcelHandler
from ValidationRules import ValidationRules
import csv
from DB import DB
from tables import table1
import openpyxl as xl

InputFileName = "testFile.xlsx"

# TASK 1 READ THE EXCEL FILE:

excelHandler = ExcelHandler(fileName=InputFileName, dataOnlyFlag=True)

rules = ValidationRules

db = DB()

row_pass = []
row_fail = []

table = db.getTable()

# DROP SHEETS AND CREATE NEW ONES
wb = xl.load_workbook('pass_fail.xlsx')
wb.remove(wb["pass"])
wb.remove(wb["fail"])
wb.create_sheet("pass")
wb.create_sheet("fail")
ws_pass = wb["pass"]
ws_fail = wb["fail"]

# LOOP THROUGH ROWS
prev = None
for i in table:
    if rules.check_notnull(i[0]) and rules.check_type(i[0]) and rules.check_lang(i[0], "ar"):
        ws_pass.append(i)

    else:
        ws_fail.append(i)

    prev = i

wb.save(filename='pass_fail.xlsx')

db.closeConnection()
