import cx_Oracle
import pandas as pd
import config

from ValidationRules import ValidationRules
from tableChecks2 import Table2
from openpyxl import load_workbook

########################################################################################################################

vals = [['CL_AGE_GROUP_EN_V1', None, '15 - 24'],
        ['CL_AGE_GROUP_EN_V1', None, '25 - 29'],
        ['CL_AGE_GROUP_EN_V1', None, '30 - 34'],
        ['CL_AGE_GROUP_EN_V1', None, '35 - 39'],
        ['CL_AGE_GROUP_EN_V1', None, '40 and Above'],
        ['CL_AGE_GROUP_EN_V1', None, 'Total'],
        ['CL_AGE_GROUP_AR_V1', None, '15 - 24'],
        ['CL_AGE_GROUP_AR_V1', None, '25 - 29'],
        ['CL_AGE_GROUP_AR_V1', None, '30 - 34'],
        ['CL_AGE_GROUP_AR_V1', None, '35 - 39'],
        ['CL_AGE_GROUP_AR_V1', None, '40 فـأعلـــى'],
        ['CL_AGE_GROUP_AR_V1', None, 'الجملة '],
        ['CL_SEX_AR_V1', None, 'ذكور'],
        ['CL_SEX_AR_V1', None, 'إناث'],
        ['CL_SEX_AR_V1', None, 'الجملة'],
        ['CL_SEX_EN_V2', None, 'Male'],
        ['CL_SEX_EN_V2', None, 'Female'],
        ['CL_SEX_EN_V2', None, 'Total']]

ref_dict = pd.DataFrame(vals, columns=['CL_ID', 'ID', 'DESCRIPTION'])

# SETUP SAVE TO EXCEL

fileName = 'Output.xlsx'

book = load_workbook(fileName)
writer = pd.ExcelWriter(fileName, engine='openpyxl')
writer.book = book
writer.sheets = {ws.title: ws for ws in book.worksheets}

rules = ValidationRules()

########################################################################################################################

# GET TABLE FROM DB INTO PANDAS DATAFRAME

connection = cx_Oracle.connect('{0}/{1}@{2}:{3}/{4}'.format(config.username,
                                                            config.password,
                                                            config.dsn,
                                                            config.port,
                                                            config.SERVICE_NAME))

query = """SELECT * FROM relational_db_t22v2"""
df_input = pd.read_sql(query, con=connection)
connection.close()

rules2 = Table2()

df_fail = pd.DataFrame(columns=[*df_input.columns] + ['isnull', 'wrong type', 'wrong language', 'out of range lookup'])
df_pass = pd.DataFrame(columns=df_input.columns)

for column, row in df_input.iterrows():

    error_dict = rules2.table_rules(input=row, columns=df_input.columns, lookups=ref_dict)
    if len([1 for x in error_dict if len(error_dict[x]) != 0]) != 0:
        inrow = [*row.values] + [*error_dict.values()]
        df_fail = df_fail.append(pd.DataFrame([inrow], columns=df_fail.columns), ignore_index=True)

    else:
        df_pass = df_pass.append(pd.DataFrame([[*row.values]], columns=df_pass.columns), ignore_index=True)

# OUTPUT GOOD AND BAD ROWS
sheetName = 'fail'
df_fail.to_excel(writer, sheet_name=sheetName, startrow=writer.sheets[sheetName].max_row, index=False)
sheetName = 'pass'
df_pass.to_excel(writer, sheet_name=sheetName, startrow=writer.sheets[sheetName].max_row, index=False)

#############################################REPORTING###########################################################


input = df_input.assign(Obs_toNumber=pd.to_numeric(df_input['OBS_VALUE'], errors='coerce'))
input = input.assign(MONTH=[i.split('-')[1].strip() for i in input['TIME_PERIOD_M']])
input = input.assign(DATE=pd.to_datetime(input['MONTH'] + '-' + input['TIME_PERIOD_Y'].astype(str)))

# PREPROCESS DATA

for i in [*input.columns]:
    try:
        if i in ['PUBLICATION_DATE_EN']:
            input[i] = pd.to_datetime(input[i])

        elif i in ['TIME_PERIOD_Y']:
            input[i] = pd.to_numeric(input[i])
        elif i in ['TIME_PERIOD_M']:
            input[i] = [i.split('-')[1].strip() for i in input[i]]
        else:
            print()
            input[i] = input[i].str.strip()
    except:
        continue
for t in input.itertuples():
    if t.Obs_toNumber != t.Obs_toNumber:

        input.at[t.Index, 'Obs_toNumber'] = pd.to_numeric(t.OBS_VALUE[:-1])


# GET MIN MAX

temp_list = [input.Obs_toNumber.idxmin(), input.Obs_toNumber.idxmax()]
min_max = input.iloc[temp_list].sort_values(by='Obs_toNumber')

sheetName = 'min_max'
min_max.to_excel(writer, sheet_name=sheetName, startrow=writer.sheets[sheetName].max_row, index=False)

# GET DIFFERENCE AND PERCENTAGE DIFFERENCE

diff = input[['CL_AGE_GROUP_EN_V1', 'CL_SEX_EN_V2', 'DATE', 'Obs_toNumber']].sort_values(
    by=['CL_AGE_GROUP_EN_V1', 'CL_SEX_EN_V2', 'DATE']).reset_index()
diff = diff.assign(difference=None, perc_diff=None)
for i in range(len(diff) - 1):
    if diff['DATE'][i + 1].month - diff['DATE'][i].month == 1:
        diff.at[i + 1, 'difference'] = diff['Obs_toNumber'][i + 1] - diff['Obs_toNumber'][i]
        if diff['Obs_toNumber'][i] == 0:
            continue
        else:
            diff.at[i + 1, 'perc_diff'] = (diff['Obs_toNumber'][i + 1] - diff['Obs_toNumber'][i]) / \
                                          diff['Obs_toNumber'][i] * 100
    else:
        continue
diff = diff.sort_values(by=['DATE', 'CL_SEX_EN_V2', 'CL_AGE_GROUP_EN_V1']).reset_index()

sheetName = 'changes'
diff.to_excel(writer, sheet_name=sheetName, startrow=writer.sheets[sheetName].max_row, index=False)

# FREQ CHECK

freq = input.drop_duplicates(['DATE'])[['DATE']].sort_values(by='DATE').reset_index(drop=True)
freq = freq.assign(FREQ=None)
for i in range(len(freq) - 1):
    freq.at[i + 1, 'FREQ'] = freq['DATE'][i + 1].month - freq['DATE'][i].month

sheetName = 'frequency'
freq.to_excel(writer, sheet_name=sheetName, startrow=writer.sheets[sheetName].max_row, index=False)

# GET TOTALS REPORT

reported_totals = input[input['CL_AGE_GROUP_EN_V1'] == 'Total'].groupby(
    ['TIME_PERIOD_Y', 'TIME_PERIOD_M', 'CL_SEX_EN_V2']
    ).agg({"Obs_toNumber": "sum"})
actual_totals = input[input['CL_AGE_GROUP_EN_V1'] != 'Total'].groupby(['TIME_PERIOD_Y', 'TIME_PERIOD_M', 'CL_SEX_EN_V2']
                                                                      ).agg({"Obs_toNumber": "sum"})
totals = reported_totals.merge(actual_totals, on=['TIME_PERIOD_Y', 'TIME_PERIOD_M', 'CL_SEX_EN_V2'], how='right'
                               ).merge(reported_totals - actual_totals,
                                       on=['TIME_PERIOD_Y', 'TIME_PERIOD_M', 'CL_SEX_EN_V2'], how='left')
totals.columns = ['Reported Total', 'Actual Total', 'Reported-Actual']

sheetName = 'total'
totals.to_excel(writer, sheet_name=sheetName, startrow=writer.sheets[sheetName].max_row, index=False)

writer.save()
