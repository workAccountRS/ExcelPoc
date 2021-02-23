import os
from openpyxl import load_workbook


class ExcelHandler:


    def __init__(self, fileName="testFile.xlsx", dataOnlyFlag=True):
        # TO FIND THE FILE IN THE PREVIOUS LEVEL OR 2 LEVELS BEHIND
        try:
            directory = os.path.abspath('.')
        except:
            directory = os.path.abspath('..')

        filepathRead = directory + "\\Files\\"  # NAVIGATE TO THE 'Files' FOLDER WHICH CONTAINS THE file.xlsx
        filepathRead = filepathRead + fileName  # COMBINE THE FILE PATH WITH THE FILE NAME

        # self.wb = load_workbook(filepathRead)  # LOAD THE EXCEL FILE AND STORE IT IN THE wb OBJECT

        self.wb = load_workbook(fileName, data_only=dataOnlyFlag)  # LOAD THE EXCEL FILE AND STORE IT IN THE wb OBJECT



    # RETURNS ALL CELLS IN A GIVEN COLUMN
    def getColumnDataFromSheet(self, sheet="S2T Mapping", column=1):
        sheet = self.wb[sheet]
        columnData = []
        for i in range(1, sheet.max_row + 1):
            columnData.append(sheet['{0}{1}'.format(column, i)].value)
        return columnData

    # RETURNS ALL CELLS IN A GIVEN ROW
    def getRowDataFromSheet(self, sheet="S2T Mapping", row=1):
        sheet = self.wb[sheet]
        rowData = []
        letters = [chr(i) for i in range(ord('A'), ord('Z') + 1)]
        for i in range(0, sheet.max_column):
            rowData.append(sheet['{0}{1}'.format(letters[i], row)].value)
        return rowData

    def getCellFromSheet(self, sheet="S2T Mapping", cell="A1"):
        sheet = self.wb[sheet]
        return sheet[str(cell)].value

    def getMaxRow(self, sheet="S2T Mapping", cell="A1"):
        sheet = self.wb[sheet]
        return sheet.max_row

    def getMaxColumn(self, sheet="S2T Mapping", cell="A1"):
        sheet = self.wb[sheet]
        return sheet.max_column

    def writeCell(self, sheet="S2T Mapping", cell="A1", value=0):
        sheet = self.wb[sheet]
        sheet[cell] = value

    def saveSpreadSheet(self, fileName='testFile.xlsx'):
        self.wb.save(filename=fileName)
